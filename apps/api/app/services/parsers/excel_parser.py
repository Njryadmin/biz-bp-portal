"""
apps/api/app/services/parsers/excel_parser.py

openpyxl-based Excel parser. Returns a list of plain dicts so the
caller can JSON-serialize and persist into `raw.uploads.payload` (jsonb).

Public API:
    parse_excel(source, sheet_name=None) -> list[dict]

Args:
    source: filesystem path (``str`` / ``Path``) or raw file ``bytes``.
    sheet_name: optional sheet name. Defaults to the active sheet.

Behaviour:
    * First non-empty row is treated as the header.
    * Empty rows are skipped.
    * Cell values are coerced to JSON-safe types (str / int / float /
      bool / isoformat-date).
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

import openpyxl


def _coerce_cell(v: Any) -> Any:
    """Make a cell value JSON-serializable."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float, str)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (list, tuple)):
        return [_coerce_cell(x) for x in v]
    return str(v)


def parse_excel(
    source: str | bytes | Path,
    sheet_name: str | None = None,
) -> list[dict[str, Any]]:
    """Parse an .xlsx / .xlsm file into a list of dicts.

    See module docstring for details.
    """
    if isinstance(source, bytes):
        wb = openpyxl.load_workbook(BytesIO(source), data_only=True, read_only=True)
    else:
        wb = openpyxl.load_workbook(str(source), data_only=True, read_only=True)

    ws = wb[sheet_name] if sheet_name else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            name = f"col_{i}"
        else:
            name = str(cell).strip()
        # Disambiguate duplicate column names so the resulting dict keys are unique.
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    out: list[dict[str, Any]] = []
    for row in rows_iter:
        # Skip completely-empty rows.
        if all(c is None for c in row):
            continue
        record: dict[str, Any] = {}
        for h, v in zip(headers, row):
            record[h] = _coerce_cell(v)
        out.append(record)
    return out

"""
apps/api/app/services/parsers/tests/test_parsers.py

Unit tests for the file parsers used by the data-integration layer.

Run from apps/api with PYTHONPATH set to its CWD:
    cd apps/api
    $env:PYTHONPATH="$PWD"
    python -m pytest app/services/parsers/tests/ -v
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook


# ---- Excel parser -------------------------------------------------------


def test_excel_parser_parses_rows(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["project_id", "date", "sales_amount"])
    ws.append(["P001", "2026-01-01", 1_500_000])
    ws.append(["P002", "2026-01-02", 1_800_000])
    path = tmp_path / "test.xlsx"
    wb.save(path)

    from app.services.parsers.excel_parser import parse_excel

    rows = parse_excel(str(path))
    assert len(rows) == 2
    assert rows[0] == {
        "project_id": "P001",
        "date": "2026-01-01",
        "sales_amount": 1_500_000,
    }
    assert rows[1]["sales_amount"] == 1_800_000


def test_excel_parser_accepts_bytes(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["k", "v"])
    ws.append(["a", 1])
    path = tmp_path / "in_memory.xlsx"
    wb.save(path)
    blob = path.read_bytes()

    from app.services.parsers.excel_parser import parse_excel

    rows = parse_excel(blob)
    assert rows == [{"k": "a", "v": 1}]


def test_excel_parser_handles_empty_file(tmp_path: Path) -> None:
    wb = Workbook()  # default empty sheet
    path = tmp_path / "empty.xlsx"
    wb.save(path)

    from app.services.parsers.excel_parser import parse_excel

    assert parse_excel(str(path)) == []


def test_excel_parser_serializes_to_json(tmp_path: Path) -> None:
    """Each returned dict must be JSON-serializable."""
    wb = Workbook()
    ws = wb.active
    ws.append(["name", "score", "active"])
    ws.append(["alice", 9.5, True])
    path = tmp_path / "j.xlsx"
    wb.save(path)

    from app.services.parsers.excel_parser import parse_excel

    rows = parse_excel(str(path))
    # Will raise TypeError if any value is not JSON-serializable.
    payload = json.dumps(rows, ensure_ascii=False)
    assert json.loads(payload) == rows


# ---- CSV parser ---------------------------------------------------------


def test_csv_parser_basic(tmp_path: Path) -> None:
    p = tmp_path / "test.csv"
    p.write_text(
        "project_id,date,sales_amount\nP001,2026-01-01,1500000\nP002,2026-01-02,1800000\n",
        encoding="utf-8",
    )

    from app.services.parsers.csv_parser import parse_csv

    rows = parse_csv(p)
    assert len(rows) == 2
    assert rows[0]["project_id"] == "P001"
    assert rows[0]["sales_amount"] == 1_500_000
    assert rows[1]["date"] == "2026-01-02"


def test_csv_parser_with_utf8_bom(tmp_path: Path) -> None:
    p = tmp_path / "bom.csv"
    p.write_bytes(b"\xef\xbb\xbfproject_id,date\nP001,2026-01-01\n")

    from app.services.parsers.csv_parser import parse_csv

    rows = parse_csv(p)
    assert rows[0]["project_id"] == "P001"  # BOM must be stripped


def test_csv_parser_empty_cells_become_none(tmp_path: Path) -> None:
    p = tmp_path / "empty_cells.csv"
    p.write_text("a,b,c\n1,,3\n", encoding="utf-8")

    from app.services.parsers.csv_parser import parse_csv

    rows = parse_csv(p)
    assert rows == [{"a": 1, "b": None, "c": 3}]


def test_csv_parser_accepts_bytes(tmp_path: Path) -> None:
    p = tmp_path / "raw.csv"
    p.write_text("k,v\nx,1\n", encoding="utf-8")
    blob = p.read_bytes()

    from app.services.parsers.csv_parser import parse_csv

    rows = parse_csv(blob)
    assert rows == [{"k": "x", "v": 1}]


# ---- Bank-statement parser ---------------------------------------------


_ICBC_SAMPLE = """中国工商银行 账户历史明细
账号: 6222 0202 0000 1234567
查询区间: 2026-01-01 至 2026-01-31

2026-01-05    09:23:15    工资收入          15000.00      18500.50
2026-01-06    14:32:00    支付宝消费-餐饮    128.50      18372.00
2026-01-08    11:00:00    ATM取现           2000.00      16372.00
2026-01-12    08:15:00    水电费代扣          256.30      21115.70
2026-01-15    10:30:00    工资收入          15000.00      36115.70
2026-01-22    14:00:00    退款                50.00      30706.90
2026-01-25    09:30:00    投资理财赎回      10000.00      40706.90
2026-01-28    15:00:00    信用卡还款         3500.00      37206.90
"""


def test_bank_statement_icbc(tmp_path: Path) -> None:
    p = tmp_path / "icbc.txt"
    p.write_text(_ICBC_SAMPLE, encoding="utf-8")

    from app.services.parsers.bank_statement import parse_bank_statement

    rows = parse_bank_statement(p)
    assert len(rows) == 8
    assert rows[0]["date"] == "2026-01-05"
    assert rows[0]["time"] == "09:23:15"
    assert rows[0]["description"] == "工资收入"
    assert rows[0]["amount"] == 15_000.00
    assert rows[0]["balance"] == 18_500.50
    assert rows[0]["direction"] == "in"
    assert rows[0]["bank"] == "ICBC"

    # 消费/取现/代扣/还款 → out
    for i in (1, 2, 3, 7):
        assert rows[i]["direction"] == "out", rows[i]
    # 工资收入/退款/赎回 → in
    for i in (0, 5, 6):
        assert rows[i]["direction"] == "in", rows[i]


_CMB_SAMPLE = """招商银行 对账单
账号: 6225 8888 8888 1234
2026/01/05\t工资\t+15000.00\t18500.50
2026/01/06\t支付宝\t-128.50\t18372.00
2026/01/08\tATM\t-2000.00\t16372.00
2026/01/15\t退款\t+50.00\t16422.00
"""


def test_bank_statement_cmb(tmp_path: Path) -> None:
    p = tmp_path / "cmb.txt"
    p.write_text(_CMB_SAMPLE, encoding="utf-8")

    from app.services.parsers.bank_statement import parse_bank_statement

    rows = parse_bank_statement(p)
    assert len(rows) == 4
    assert rows[0]["date"] == "2026-01-05"
    assert rows[0]["description"] == "工资"
    assert rows[0]["amount"] == 15_000.00
    assert rows[0]["balance"] == 18_500.50
    assert rows[0]["direction"] == "in"
    assert rows[0]["bank"] == "CMB"
    assert rows[1]["direction"] == "out"
    assert rows[1]["amount"] == 128.50
    assert rows[3]["direction"] == "in"
    assert rows[3]["amount"] == 50.00


def test_bank_statement_skips_non_transaction_lines(tmp_path: Path) -> None:
    p = tmp_path / "mixed.txt"
    p.write_text(
        "中国工商银行 账户明细\n"
        "some random header line that is not a transaction\n"
        "2026-01-05 09:23:15 工资收入 15000.00 余额 18500.50\n"
        "another random line\n"
        "2026-01-06 14:32:00 消费 128.50 余额 18372.00\n",
        encoding="utf-8",
    )

    from app.services.parsers.bank_statement import parse_bank_statement

    rows = parse_bank_statement(p)
    assert len(rows) == 2
    assert [r["description"] for r in rows] == ["工资收入", "消费"]


def test_bank_statement_accepts_bytes() -> None:
    blob = _ICBC_SAMPLE.encode("utf-8")

    from app.services.parsers.bank_statement import parse_bank_statement

    rows = parse_bank_statement(blob)
    assert len(rows) >= 5
    assert all("date" in r and "amount" in r and "balance" in r for r in rows)
